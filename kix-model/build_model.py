"""Build the KIX venue financial model workbook."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------- styling ----
FONT = "Arial"

TITLE = Font(name=FONT, size=16, bold=True, color="1A1A1A")
H1 = Font(name=FONT, size=12, bold=True, color="FFFFFF")
H2 = Font(name=FONT, size=10, bold=True, color="1A1A1A")
BODY = Font(name=FONT, size=10)
BODY_B = Font(name=FONT, size=10, bold=True)
INPUT = Font(name=FONT, size=10, color="0000FF")          # hardcoded input
LINK = Font(name=FONT, size=10, color="008000")           # cross-sheet link
NOTE = Font(name=FONT, size=9, italic=True, color="595959")
TOTAL = Font(name=FONT, size=10, bold=True, color="1A1A1A")

BAND = PatternFill("solid", fgColor="1F3864")             # section header band
KEY = PatternFill("solid", fgColor="FFFF00")              # key assumption
SUB = PatternFill("solid", fgColor="D9E2F3")              # subtotal shading
WARN = PatternFill("solid", fgColor="FCE4D6")

THIN = Side(style="thin", color="BFBFBF")
TOPLINE = Border(top=Side(style="thin", color="404040"))
BOXED = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

MONEY = '$#,##0;($#,##0);-'
MONEY2 = '$#,##0.00;($#,##0.00);-'
NUM = '#,##0;(#,##0);-'
NUM1 = '#,##0.0;(#,##0.0);-'
PCT = '0.0%;(0.0%);-'
MULT = '0.00x'
YRS = '0.0'


def band(ws, row, text, last_col=4):
    """Full-width section header."""
    ws.cell(row=row, column=1, value=text).font = H1
    for c in range(1, last_col + 1):
        ws.cell(row=row, column=c).fill = BAND
    ws.row_dimensions[row].height = 18


def label(ws, row, text, bold=False, indent=0):
    c = ws.cell(row=row, column=1, value=text)
    c.font = BODY_B if bold else BODY
    if indent:
        c.alignment = Alignment(indent=indent)
    return c


def put(ws, row, col, value, font=BODY, fmt=None, fill=None, border=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font
    if fmt:
        c.number_format = fmt
    if fill:
        c.fill = fill
    if border:
        c.border = border
    return c


def note(ws, row, col, text):
    c = ws.cell(row=row, column=col, value=text)
    c.font = NOTE
    return c


wb = Workbook()
# LibreOffice is unavailable in this environment, so no cached values can be written.
# Force Excel / Sheets / Numbers to compute every formula the moment the file opens.

# =============================================================== README ======
ws = wb.active
ws.title = "README"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 108

r = 2
put(ws, r, 2, "KIX — VENUE FINANCIAL MODEL", TITLE); r += 1
put(ws, r, 2, "Bottom-up capacity model - 18,000 SF vs 35,000 SF - benchmarked to a real ESA-equipped venue",
    Font(name=FONT, size=11, color="595959")); r += 2

readme = [
    ("H", "What this model does"),
    ("P", "Builds venue revenue from the bottom up - stations x throughput x operating hours x utilisation - "
          "instead of assuming a revenue target. Every line traces back to a driver you can change."),
    ("P", "It runs two footprints side by side: an 18,000 SF box and the 35,000 SF box in the current KIX plan, "
          "so the trade-off between size, capital and payback is explicit."),
    ("P", "It is calibrated against Ballpark Brighton - the only published financial case study of an "
          "ESA-equipped entertainment venue - so the assumptions are anchored to a real operating comparable."),
    ("S", ""),
    ("H", "How to use it"),
    ("P", "1.  Open 'Assumptions'. Every blue cell is an input. Change them; everything else recalculates."),
    ("P", "2.  'Revenue Build' shows the stabilised (Year 3) revenue for each footprint and cross-checks it "
          "against the benchmark's revenue density."),
    ("P", "3.  'Capex' builds total project cost excluding real estate."),
    ("P", "4.  'P&L' applies the ramp to produce five years for each footprint."),
    ("P", "5.  'Returns & Scenarios' gives payback, cash-on-cash, downside/base/upside, and a direct "
          "reconciliation against the numbers in the current KIX business plan."),
    ("S", ""),
    ("H", "Colour legend"),
    ("B", "Blue text        = hardcoded input. These are the cells to edit."),
    ("G", "Green text       = link to another sheet in this workbook."),
    ("P", "Black text       = calculated. Do not overwrite."),
    ("Y", "Yellow fill      = key assumption. These drive the result more than anything else."),
    ("S", ""),
    ("H", "Health warning"),
    ("P", "Two assumptions carry almost all the risk in this model: blended station utilisation and spend per "
          "head. Both are guesses until Sarasota (or a presell campaign) produces real data. Everything else is "
          "arithmetic. Flex those two first and see how fast the returns move."),
    ("P", "Nothing in here is validated demand. It is a planning case, not a forecast."),
    ("S", ""),
    ("H", "Sources"),
    ("P", "Benchmark data: Elite Skills Arena, 'Ballpark Brighton: Financial Performance of an ESA-Based "
          "Entertainment Venue', published 18 May 2026."),
    ("P", "https://eliteskillsarena.com/post/ballpark-brighton-financial-performance-of-an-esa-based-entertainment-venue"),
    ("P", "KIX plan figures: KIX Master Business Plan 2026, section 26 (Unit Economics) and section 9 (buildout target)."),
    ("P", "All other assumptions are management estimates entered as inputs and flagged blue in 'Assumptions'."),
]

for kind, text in readme:
    if kind == "S":
        r += 1
        continue
    c = ws.cell(row=r, column=2, value=text)
    if kind == "H":
        c.font = Font(name=FONT, size=11, bold=True, color="1F3864")
    elif kind == "B":
        c.font = INPUT
    elif kind == "G":
        c.font = LINK
    elif kind == "Y":
        c.font = BODY
        c.fill = KEY
    else:
        c.font = BODY
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r].height = 28 if len(text) > 100 else 14
    r += 1

# ============================================================ BENCHMARK ======
ws = wb.create_sheet("Benchmark")
ws.sheet_view.showGridLines = False
widths = {"A": 46, "B": 16, "C": 16, "D": 52}
for k, v in widths.items():
    ws.column_dimensions[k].width = v

r = 1
put(ws, r, 1, "BENCHMARK — BALLPARK BRIGHTON", TITLE); r += 1
put(ws, r, 1, "The only published financial case study of an ESA-equipped entertainment venue. "
              "Source: Elite Skills Arena, 18 May 2026.", NOTE); r += 2

band(ws, r, "FX AND REPORTED FIGURES"); r += 1
put(ws, r, 2, "GBP", H2); put(ws, r, 3, "USD", H2); put(ws, r, 4, "Note", H2); r += 1

FX_ROW = r
label(ws, r, "GBP / USD rate")
put(ws, r, 2, 1.30, INPUT, '0.00', KEY)
note(ws, r, 4, "Edit to change every USD figure below")
r += 1

BP_SF = r
label(ws, r, "Footprint (sq ft)")
put(ws, r, 2, 5382, INPUT, NUM)
note(ws, r, 4, "500 sq m as reported, converted at 10.764 sq ft per sq m")
r += 1

BP_REV = r
label(ws, r, "Annual revenue")
put(ws, r, 2, 192000, INPUT, MONEY)
put(ws, r, 3, f"=B{BP_REV}*$B${FX_ROW}", BODY, MONEY)
note(ws, r, 4, "GBP 16,000 per month as reported")
r += 1

BP_EQ = r
label(ws, r, "ESA equipment cost")
put(ws, r, 2, 69000, INPUT, MONEY)
put(ws, r, 3, f"=B{BP_EQ}*$B${FX_ROW}", BODY, MONEY)
note(ws, r, 4, "Fast Feet, ICON 6m, Precision Wall")
r += 1

BP_CAP = r
label(ws, r, "Total opening cost")
put(ws, r, 2, 200000, INPUT, MONEY)
put(ws, r, 3, f"=B{BP_CAP}*$B${FX_ROW}", BODY, MONEY)
note(ws, r, 4, "Converted restaurant, not a raw warehouse shell")
r += 1

BP_RENT = r
label(ws, r, "Annual rent")
put(ws, r, 2, 45000, INPUT, MONEY)
put(ws, r, 3, f"=B{BP_RENT}*$B${FX_ROW}", BODY, MONEY)
note(ws, r, 4, "GBP 3,750 per month as reported")
r += 1

BP_FB = r
label(ws, r, "Bar / F&B share of revenue")
put(ws, r, 2, 0.15, INPUT, PCT)
note(ws, r, 4, "Low. A proper F&B operation runs 25-40% in this category")
r += 2

band(ws, r, "DERIVED DENSITY"); r += 1
BP_PSF = r
label(ws, r, "Revenue per SF per year", bold=True)
put(ws, r, 2, f"=B{BP_REV}/B{BP_SF}", TOTAL, MONEY2)
put(ws, r, 3, f"=C{BP_REV}/B{BP_SF}", TOTAL, MONEY2, KEY)
note(ws, r, 4, "THE anchor number for this whole model")
r += 1

BP_EQ_PSF = r
label(ws, r, "Equipment cost per SF")
put(ws, r, 2, f"=B{BP_EQ}/B{BP_SF}", BODY, MONEY2)
put(ws, r, 3, f"=C{BP_EQ}/B{BP_SF}", BODY, MONEY2)
r += 1

BP_CAP_PSF = r
label(ws, r, "Total capex per SF")
put(ws, r, 2, f"=B{BP_CAP}/B{BP_SF}", BODY, MONEY2)
put(ws, r, 3, f"=C{BP_CAP}/B{BP_SF}", BODY, MONEY2)
note(ws, r, 4, "Fitted restaurant conversion. A US warehouse shell costs materially more")
r += 1

BP_EQ_PCT = r
label(ws, r, "Equipment as % of total capex")
put(ws, r, 2, f"=B{BP_EQ}/B{BP_CAP}", BODY, PCT)
note(ws, r, 4, "34.5% as reported — useful sense-check on the KIX capex mix")
r += 1

BP_RENT_PSF = r
label(ws, r, "Rent per SF per year")
put(ws, r, 2, f"=B{BP_RENT}/B{BP_SF}", BODY, MONEY2)
put(ws, r, 3, f"=C{BP_RENT}/B{BP_SF}", BODY, MONEY2)
r += 2

band(ws, r, "THE KIX PLAN, AS WRITTEN"); r += 1
KIX_SF = r
label(ws, r, "Planned footprint (SF)")
put(ws, r, 2, 35000, INPUT, NUM)
note(ws, r, 4, "Midpoint of the 30-40k SF target, plan section 9")
r += 1
KIX_REV = r
label(ws, r, "Planned stabilised revenue (Year 3)")
put(ws, r, 2, 5000000, INPUT, MONEY)
note(ws, r, 4, "Plan section 26")
r += 1
KIX_EBITDA = r
label(ws, r, "Planned Year 3 EBITDA")
put(ws, r, 2, 1300000, INPUT, MONEY)
note(ws, r, 4, "Plan section 26 — 26% margin")
r += 1
KIX_CAPEX = r
label(ws, r, "Planned buildout (excl. real estate)")
put(ws, r, 2, 1500000, INPUT, MONEY)
note(ws, r, 4, "Midpoint of the $1.0-2.0M target, plan section 9")
r += 1
KIX_PSF = r
label(ws, r, "Implied revenue per SF", bold=True)
put(ws, r, 2, f"=B{KIX_REV}/B{KIX_SF}", TOTAL, MONEY2, WARN)
r += 1
KIX_MULT = r
label(ws, r, "Multiple of Ballpark density", bold=True)
put(ws, r, 2, f"=B{KIX_PSF}/C{BP_PSF}", TOTAL, MULT, WARN)
note(ws, r, 4, "The plan assumes this many times the revenue density of the benchmark")
r += 1
KIX_PAY = r
label(ws, r, "Implied simple payback (years)")
put(ws, r, 2, f"=B{KIX_CAPEX}/B{KIX_EBITDA}", TOTAL, YRS, WARN)
note(ws, r, 4, "A payback this fast is the number investors will challenge first")
r += 2

put(ws, r, 1, "Reading this sheet", H2); r += 1
for line in [
    "Ballpark is a husband-and-wife independent in a converted restaurant with no membership programme and F&B at only 15%.",
    "A professionally capitalised venue should beat it. But the KIX plan assumes it beats it by roughly 3x per square foot,",
    "in a footprint over six times larger — and larger boxes almost always earn LESS per SF, not more, because circulation,",
    "back-of-house, party rooms and seating do not sell tickets. That gap is the single biggest risk in the plan.",
]:
    put(ws, r, 1, line, BODY); r += 1

# =========================================================== ASSUMPTIONS ======
ws = wb.create_sheet("Assumptions")
ws.sheet_view.showGridLines = False
for k, v in {"A": 48, "B": 15, "C": 15, "D": 54}.items():
    ws.column_dimensions[k].width = v

r = 1
put(ws, r, 1, "ASSUMPTIONS", TITLE); r += 1
put(ws, r, 1, "Every blue cell is an input. Yellow fill marks the assumptions that move the answer most.", NOTE); r += 2

put(ws, r, 2, "Scenario A", H2); put(ws, r, 3, "Scenario B", H2); put(ws, r, 4, "Note", H2); r += 1
put(ws, r, 2, "18,000 SF", BODY_B); put(ws, r, 3, "35,000 SF", BODY_B); r += 1

A = {}  # named row registry


def arow(key, text, a_val, b_val, fmt, note_text="", key_assumption=False):
    global r
    label(ws, r, text)
    fill = KEY if key_assumption else None
    put(ws, r, 2, a_val, INPUT, fmt, fill)
    put(ws, r, 3, b_val, INPUT, fmt, fill)
    if note_text:
        note(ws, r, 4, note_text)
    A[key] = r
    r += 1


band(ws, r, "FOOTPRINT & LAYOUT"); r += 1
arow("sf", "Total footprint (SF)", 18000, 35000, NUM, "The core trade-off this model tests", True)
arow("stations", "Attraction stations", 14, 22, NUM, "ESA units: Power Shot, Precision Wall, Fast Feet, Dribble Lane, Panenka, Sprint Lane")
arow("party_rooms", "Party rooms", 2, 4, NUM)
arow("pitches", "Small-sided pitches", 1, 3, NUM, "Drives league capacity")
r += 1

band(ws, r, "CAPACITY & THROUGHPUT"); r += 1
arow("cycles_hr", "Play cycles per station per hour", 7, 7, NUM1, "Allows for changeover and group rotation")
arow("hours_wk", "Operating hours per week", 85, 85, NUM, "~12h weekdays, ~13h weekends")
arow("weeks", "Weeks open per year", 51, 51, NUM)
arow("util", "Blended station utilisation", 0.20, 0.20, PCT,
     "THE most sensitive assumption. Demand concentrates into after-school and weekends", True)
arow("cycles_visit", "Play cycles per walk-in visit", 5, 5, NUM1, "A visitor buys a game pack, not one play")
r += 1

band(ws, r, "PRICING & SPEND"); r += 1
arow("spend_visit", "Attraction spend per walk-in visit ($)", 26, 26, MONEY, "Game pack price point", True)
arow("fb_head", "F&B spend per head of footfall ($)", 6.00, 6.00, MONEY2,
     "Blended across all footfall. Ballpark achieves far less; a real F&B programme should beat this", True)
arow("retail_head", "Retail spend per head of footfall ($)", 0.90, 0.90, MONEY2)
r += 1

band(ws, r, "MEMBERSHIP"); r += 1
arow("members", "Average active members", 300, 500, NUM, "Net of churn, averaged across the year", True)
arow("member_fee", "Average monthly fee ($)", 85, 85, MONEY, "Blended across PLAY/PLAYER/PRO/FAMILY tiers")
arow("member_visits", "Member visits per member per month", 2.5, 2.5, NUM1,
     "Feeds footfall for F&B and retail. Member play is covered by the fee, not charged again")
r += 1

band(ws, r, "PROGRAMMES"); r += 1
arow("parties", "Parties per year", 300, 450, NUM, "Ballpark runs 1-2 per week in 5,400 SF")
arow("party_val", "Average party package ($)", 600, 600, MONEY)
arow("party_guests", "Average guests per party", 16, 16, NUM)
arow("academy_rev", "Academy & clinics revenue ($)", 220000, 340000, MONEY)
arow("academy_ff", "Academy participant visits (annual)", 9000, 14000, NUM)
arow("camp_rev", "Camps revenue ($)", 120000, 190000, MONEY, "School holidays and summer")
arow("camp_ff", "Camp participant-days (annual)", 3000, 4700, NUM)
arow("league_rev", "Leagues revenue ($)", 150000, 320000, MONEY, "Scales with number of pitches")
arow("league_ff", "League player-visits (annual)", 7000, 15000, NUM)
arow("corp_rev", "Corporate & private events revenue ($)", 140000, 240000, MONEY)
arow("corp_ff", "Corporate event attendees (annual)", 2800, 4800, NUM)
arow("sponsor", "Sponsorship revenue ($)", 40000, 60000, MONEY)
r += 1

band(ws, r, "CAPEX"); r += 1
arow("fitout_psf", "Fit-out cost per SF ($)", 60, 60, MONEY,
     "Warehouse shell to finished venue. Lean 55 / realistic 85 / premium 120", True)
arow("station_cost", "Cost per attraction station installed ($)", 24000, 24000, MONEY,
     "ESA list plus freight, duty, install. Benchmark implies ~$18k before US landed costs")
arow("ffe", "FF&E, bar & kitchen build ($)", 220000, 380000, MONEY)
arow("tech_capex", "Technology: KIX ID, POS, AV, network ($)", 180000, 220000, MONEY,
     "Build the identity and scoring layer; buy the venue-ops plumbing")
arow("preopen", "Pre-opening: marketing, hiring, licences ($)", 150000, 220000, MONEY)
arow("conting", "Contingency (% of subtotal)", 0.10, 0.10, PCT)
r += 1

band(ws, r, "OPERATING COSTS"); r += 1
arow("labour", "Labour (% of revenue)", 0.31, 0.30, PCT, "Includes coaches, party hosts, floor and F&B staff", True)
arow("rent_psf", "Rent per SF per year ($)", 14, 13, MONEY, "Florida warehouse / flex, NNN")
arow("util_psf", "Utilities per SF per year ($)", 3.50, 3.50, MONEY2)
arow("ins_psf", "Insurance per SF per year ($)", 2.50, 2.50, MONEY2,
     "Florida property and liability. Materially higher than most US states")
arow("mktg", "Marketing (% of revenue)", 0.06, 0.06, PCT)
arow("fb_cogs", "F&B COGS (% of F&B revenue)", 0.28, 0.28, PCT)
arow("svc", "Equipment service (% of equipment capex)", 0.08, 0.08, PCT, "Spares, callouts, consumables")
arow("software", "Software & systems ($ per year)", 60000, 75000, MONEY)
arow("repairs", "Repairs & renewals (% of revenue)", 0.02, 0.02, PCT)
arow("admin", "Admin & other (% of revenue)", 0.05, 0.05, PCT)
r += 1

band(ws, r, "REVENUE RAMP (% of stabilised)"); r += 1
arow("ramp1", "Year 1", 0.55, 0.55, PCT, "A new venue with no brand does not open at stabilised volume")
arow("ramp2", "Year 2", 0.82, 0.82, PCT)
arow("ramp3", "Year 3 — stabilised", 1.00, 1.00, PCT)
arow("ramp4", "Year 4", 1.06, 1.06, PCT)
arow("ramp5", "Year 5", 1.11, 1.11, PCT)


def AS(key, col):
    return f"Assumptions!${col}${A[key]}"


# ========================================================= REVENUE BUILD ======
ws = wb.create_sheet("Revenue Build")
ws.sheet_view.showGridLines = False
for k, v in {"A": 46, "B": 16, "C": 16, "D": 52}.items():
    ws.column_dimensions[k].width = v

r = 1
put(ws, r, 1, "REVENUE BUILD — STABILISED (YEAR 3)", TITLE); r += 1
put(ws, r, 1, "Built from capacity, not from a target.", NOTE); r += 2
put(ws, r, 2, "18,000 SF", H2); put(ws, r, 3, "35,000 SF", H2); put(ws, r, 4, "Note", H2); r += 1

R = {}


def rrow(key, text, fa, fb, fmt, note_text="", bold=False, fill=None, top=False):
    global r
    label(ws, r, text, bold=bold)
    bd = TOPLINE if top else None
    put(ws, r, 2, fa, TOTAL if bold else BODY, fmt, fill, bd)
    put(ws, r, 3, fb, TOTAL if bold else BODY, fmt, fill, bd)
    if note_text:
        note(ws, r, 4, note_text)
    R[key] = r
    r += 1


band(ws, r, "CAPACITY"); r += 1
rrow("cap_hr", "Play cycles available per hour",
     f"={AS('stations','B')}*{AS('cycles_hr','B')}",
     f"={AS('stations','C')}*{AS('cycles_hr','C')}", NUM)
rrow("cap_yr", "Play cycles available per year",
     f"=B{R['cap_hr']}*{AS('hours_wk','B')}*{AS('weeks','B')}",
     f"=C{R['cap_hr']}*{AS('hours_wk','C')}*{AS('weeks','C')}", NUM,
     "Theoretical maximum at 100% utilisation")
rrow("cap_used", "Play cycles sold",
     f"=B{R['cap_yr']}*{AS('util','B')}",
     f"=C{R['cap_yr']}*{AS('util','C')}", NUM)
rrow("walkins", "Walk-in visits per year",
     f"=B{R['cap_used']}/{AS('cycles_visit','B')}",
     f"=C{R['cap_used']}/{AS('cycles_visit','C')}", NUM, "", bold=True)
r += 1

band(ws, r, "FOOTFALL (drives F&B and retail)"); r += 1
rrow("ff_walk", "Walk-in visitors", f"=B{R['walkins']}", f"=C{R['walkins']}", NUM)
rrow("ff_mem", "Member visits",
     f"={AS('members','B')}*{AS('member_visits','B')}*12",
     f"={AS('members','C')}*{AS('member_visits','C')}*12", NUM)
rrow("ff_party", "Party guests",
     f"={AS('parties','B')}*{AS('party_guests','B')}",
     f"={AS('parties','C')}*{AS('party_guests','C')}", NUM)
rrow("ff_acad", "Academy participants", f"={AS('academy_ff','B')}", f"={AS('academy_ff','C')}", NUM)
rrow("ff_camp", "Camp participant-days", f"={AS('camp_ff','B')}", f"={AS('camp_ff','C')}", NUM)
rrow("ff_league", "League player-visits", f"={AS('league_ff','B')}", f"={AS('league_ff','C')}", NUM)
rrow("ff_corp", "Corporate attendees", f"={AS('corp_ff','B')}", f"={AS('corp_ff','C')}", NUM)
rrow("ff_total", "Total annual footfall",
     f"=SUM(B{R['ff_walk']}:B{R['ff_corp']})",
     f"=SUM(C{R['ff_walk']}:C{R['ff_corp']})", NUM, "", bold=True, fill=SUB, top=True)
r += 1

band(ws, r, "REVENUE STREAMS"); r += 1
rrow("rev_walk", "1. Walk-in pay-to-play",
     f"=B{R['ff_walk']}*{AS('spend_visit','B')}",
     f"=C{R['ff_walk']}*{AS('spend_visit','C')}", MONEY)
rrow("rev_mem", "2. Membership",
     f"={AS('members','B')}*{AS('member_fee','B')}*12",
     f"={AS('members','C')}*{AS('member_fee','C')}*12", MONEY, "Member play is covered by the fee")
rrow("rev_party", "3. Parties",
     f"={AS('parties','B')}*{AS('party_val','B')}",
     f"={AS('parties','C')}*{AS('party_val','C')}", MONEY)
rrow("rev_acad", "4. Academy & clinics", f"={AS('academy_rev','B')}", f"={AS('academy_rev','C')}", MONEY)
rrow("rev_camp", "5. Camps", f"={AS('camp_rev','B')}", f"={AS('camp_rev','C')}", MONEY)
rrow("rev_league", "6. Leagues", f"={AS('league_rev','B')}", f"={AS('league_rev','C')}", MONEY)
rrow("rev_corp", "7. Corporate & private events", f"={AS('corp_rev','B')}", f"={AS('corp_rev','C')}", MONEY)
rrow("rev_fb", "8. Food & beverage",
     f"=B{R['ff_total']}*{AS('fb_head','B')}",
     f"=C{R['ff_total']}*{AS('fb_head','C')}", MONEY, "The margin engine. Do not under-build this")
rrow("rev_retail", "9. Retail",
     f"=B{R['ff_total']}*{AS('retail_head','B')}",
     f"=C{R['ff_total']}*{AS('retail_head','C')}", MONEY)
rrow("rev_sponsor", "10. Sponsorship", f"={AS('sponsor','B')}", f"={AS('sponsor','C')}", MONEY)
rrow("rev_total", "STABILISED REVENUE",
     f"=SUM(B{R['rev_walk']}:B{R['rev_sponsor']})",
     f"=SUM(C{R['rev_walk']}:C{R['rev_sponsor']})", MONEY, "", bold=True, fill=SUB, top=True)
r += 1

band(ws, r, "MIX CHECK"); r += 1
rrow("mix_fb", "F&B as % of revenue",
     f"=B{R['rev_fb']}/B{R['rev_total']}", f"=C{R['rev_fb']}/C{R['rev_total']}", PCT,
     "Benchmark achieves 15%. Category norm is 25-40%")
rrow("mix_mem", "Membership as % of revenue",
     f"=B{R['rev_mem']}/B{R['rev_total']}", f"=C{R['rev_mem']}/C{R['rev_total']}", PCT,
     "Reality check on the 'recurring revenue platform' story")
rrow("mix_rec", "Recurring + programmed as % of revenue",
     f"=(B{R['rev_mem']}+B{R['rev_acad']}+B{R['rev_camp']}+B{R['rev_league']})/B{R['rev_total']}",
     f"=(C{R['rev_mem']}+C{R['rev_acad']}+C{R['rev_camp']}+C{R['rev_league']})/C{R['rev_total']}", PCT)
r += 1

band(ws, r, "DENSITY CROSS-CHECK"); r += 1
rrow("psf", "Revenue per SF per year",
     f"=B{R['rev_total']}/{AS('sf','B')}", f"=C{R['rev_total']}/{AS('sf','C')}", MONEY2, "", bold=True)
rrow("bench_psf", "Ballpark Brighton benchmark ($/SF)",
     f"=Benchmark!$C${BP_PSF}", f"=Benchmark!$C${BP_PSF}", MONEY2)
for cc in (2, 3):
    ws.cell(row=R['bench_psf'], column=cc).font = LINK
rrow("bench_mult", "Multiple of benchmark density",
     f"=B{R['psf']}/B{R['bench_psf']}", f"=C{R['psf']}/C{R['bench_psf']}", MULT,
     "How much better than the comparable this model assumes you operate", bold=True)
rrow("vs_plan", "vs KIX plan revenue ($5.0M)",
     f"=B{R['rev_total']}/Benchmark!$B${KIX_REV}",
     f"=C{R['rev_total']}/Benchmark!$B${KIX_REV}", PCT,
     "This model as a percentage of the plan's Year 3 number", bold=True)
r += 1

put(ws, r, 1, "Note: member attraction play is not charged twice — the membership fee covers it, and member "
              "visits appear in footfall only, where they drive F&B and retail.", NOTE)

# ================================================================ CAPEX ======
ws = wb.create_sheet("Capex")
ws.sheet_view.showGridLines = False
for k, v in {"A": 46, "B": 16, "C": 16, "D": 52}.items():
    ws.column_dimensions[k].width = v

r = 1
put(ws, r, 1, "CAPEX — TOTAL PROJECT COST", TITLE); r += 1
put(ws, r, 1, "Excludes real estate acquisition.", NOTE); r += 2
put(ws, r, 2, "18,000 SF", H2); put(ws, r, 3, "35,000 SF", H2); put(ws, r, 4, "Note", H2); r += 1

C = {}


def crow(key, text, fa, fb, fmt, note_text="", bold=False, fill=None, top=False):
    global r
    label(ws, r, text, bold=bold)
    bd = TOPLINE if top else None
    put(ws, r, 2, fa, TOTAL if bold else BODY, fmt, fill, bd)
    put(ws, r, 3, fb, TOTAL if bold else BODY, fmt, fill, bd)
    if note_text:
        note(ws, r, 4, note_text)
    C[key] = r
    r += 1


band(ws, r, "PROJECT COST"); r += 1
crow("fitout", "Fit-out (shell to finished venue)",
     f"={AS('sf','B')}*{AS('fitout_psf','B')}", f"={AS('sf','C')}*{AS('fitout_psf','C')}", MONEY,
     "HVAC, power, restrooms, fire, ADA, flooring, lighting, finishes")
crow("equip", "Attraction equipment (ESA)",
     f"={AS('stations','B')}*{AS('station_cost','B')}",
     f"={AS('stations','C')}*{AS('station_cost','C')}", MONEY, "Landed and installed")
crow("ffe", "FF&E, bar & kitchen", f"={AS('ffe','B')}", f"={AS('ffe','C')}", MONEY)
crow("tech", "Technology", f"={AS('tech_capex','B')}", f"={AS('tech_capex','C')}", MONEY)
crow("preopen", "Pre-opening", f"={AS('preopen','B')}", f"={AS('preopen','C')}", MONEY)
crow("subtotal", "Subtotal",
     f"=SUM(B{C['fitout']}:B{C['preopen']})", f"=SUM(C{C['fitout']}:C{C['preopen']})", MONEY,
     "", bold=True, top=True)
crow("conting", "Contingency",
     f"=B{C['subtotal']}*{AS('conting','B')}", f"=C{C['subtotal']}*{AS('conting','C')}", MONEY)
crow("total", "TOTAL PROJECT COST",
     f"=B{C['subtotal']}+B{C['conting']}", f"=C{C['subtotal']}+C{C['conting']}", MONEY,
     "", bold=True, fill=SUB, top=True)
r += 1

band(ws, r, "CAPEX CHECKS"); r += 1
crow("psf", "Cost per SF",
     f"=B{C['total']}/{AS('sf','B')}", f"=C{C['total']}/{AS('sf','C')}", MONEY2)
crow("bench_psf", "Ballpark benchmark cost per SF",
     f"=Benchmark!$C${BP_CAP_PSF}", f"=Benchmark!$C${BP_CAP_PSF}", MONEY2,
     "Benchmark was a fitted restaurant, so a warehouse shell should cost more")
for cc in (2, 3):
    ws.cell(row=C['bench_psf'], column=cc).font = LINK
crow("eq_pct", "Equipment as % of total capex",
     f"=B{C['equip']}/B{C['total']}", f"=C{C['equip']}/C{C['total']}", PCT,
     "Benchmark ratio is 34.5%")
crow("vs_plan", "vs KIX planned buildout ($1.5M)",
     f"=B{C['total']}/Benchmark!$B${KIX_CAPEX}", f"=C{C['total']}/Benchmark!$B${KIX_CAPEX}", MULT,
     "How far the plan's buildout budget is from a realistic build", bold=True)
for cc in (2, 3):
    ws.cell(row=C['vs_plan'], column=cc).fill = WARN

# ================================================================= P&L =======
ws = wb.create_sheet("P&L")
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 42
for i in range(2, 7):
    ws.column_dimensions[get_column_letter(i)].width = 15
ws.column_dimensions["G"].width = 44

r = 1
put(ws, r, 1, "FIVE YEAR P&L", TITLE); r += 1
put(ws, r, 1, "Stabilised revenue from 'Revenue Build', scaled by the ramp in 'Assumptions'.", NOTE); r += 2

ramp_keys = ["ramp1", "ramp2", "ramp3", "ramp4", "ramp5"]
PL = {}


def pl_block(title, col, rev_row, eq_row, sf_key_col, start_row):
    """Write one scenario's P&L. Returns dict of row indices."""
    global r
    r = start_row
    band(ws, r, title, last_col=7); r += 1
    for i in range(5):
        put(ws, r, 2 + i, f"Year {i+1}", H2)
        ws.cell(row=r, column=2 + i).alignment = Alignment(horizontal="right")
    put(ws, r, 7, "Note", H2)
    r += 1

    rows = {}

    label(ws, r, "Revenue", bold=True)
    for i in range(5):
        put(ws, r, 2 + i,
            f"='Revenue Build'!${col}${rev_row}*{AS(ramp_keys[i], col)}",
            TOTAL, MONEY)
    rows["rev"] = r; r += 1

    label(ws, r, "  of which F&B", indent=1)
    for i in range(5):
        put(ws, r, 2 + i,
            f"={get_column_letter(2+i)}{rows['rev']}*'Revenue Build'!${col}${R['mix_fb']}",
            BODY, MONEY)
    rows["fb"] = r; r += 1
    r += 1

    label(ws, r, "Operating costs", bold=True); r += 1

    cost_defs = [
        ("labour", "Labour", lambda cl: f"={cl}{rows['rev']}*{AS('labour', col)}"),
        ("fbcogs", "F&B COGS", lambda cl: f"={cl}{rows['fb']}*{AS('fb_cogs', col)}"),
        ("rent", "Rent", lambda cl: f"={AS('sf', col)}*{AS('rent_psf', col)}"),
        ("utils", "Utilities", lambda cl: f"={AS('sf', col)}*{AS('util_psf', col)}"),
        ("ins", "Insurance", lambda cl: f"={AS('sf', col)}*{AS('ins_psf', col)}"),
        ("mktg", "Marketing", lambda cl: f"={cl}{rows['rev']}*{AS('mktg', col)}"),
        ("svc", "Equipment service", lambda cl: f"=Capex!${col}${eq_row}*{AS('svc', col)}"),
        ("soft", "Software & systems", lambda cl: f"={AS('software', col)}"),
        ("rep", "Repairs & renewals", lambda cl: f"={cl}{rows['rev']}*{AS('repairs', col)}"),
        ("admin", "Admin & other", lambda cl: f"={cl}{rows['rev']}*{AS('admin', col)}"),
    ]
    first_cost = r
    for key, text, fn in cost_defs:
        label(ws, r, "  " + text, indent=1)
        for i in range(5):
            cl = get_column_letter(2 + i)
            put(ws, r, 2 + i, fn(cl), BODY, MONEY)
        rows[key] = r
        r += 1
    last_cost = r - 1

    label(ws, r, "Total operating costs", bold=True)
    for i in range(5):
        cl = get_column_letter(2 + i)
        put(ws, r, 2 + i, f"=SUM({cl}{first_cost}:{cl}{last_cost})", TOTAL, MONEY, border=TOPLINE)
    rows["totcost"] = r; r += 1
    r += 1

    label(ws, r, "EBITDA", bold=True)
    for i in range(5):
        cl = get_column_letter(2 + i)
        put(ws, r, 2 + i, f"={cl}{rows['rev']}-{cl}{rows['totcost']}", TOTAL, MONEY, SUB, TOPLINE)
    rows["ebitda"] = r; r += 1

    label(ws, r, "EBITDA margin", bold=True)
    for i in range(5):
        cl = get_column_letter(2 + i)
        put(ws, r, 2 + i, f"=IF({cl}{rows['rev']}=0,0,{cl}{rows['ebitda']}/{cl}{rows['rev']})",
            TOTAL, PCT, SUB)
    rows["margin"] = r; r += 1

    label(ws, r, "Cumulative EBITDA")
    put(ws, r, 2, f"=B{rows['ebitda']}", BODY, MONEY)
    for i in range(1, 5):
        cl = get_column_letter(2 + i)
        pv = get_column_letter(1 + i)
        put(ws, r, 2 + i, f"={pv}{r}+{cl}{rows['ebitda']}", BODY, MONEY)
    rows["cum"] = r; r += 1

    return rows


PL["A"] = pl_block("SCENARIO A — 18,000 SF", "B", R["rev_total"], C["equip"], "B", 5)
PL["B"] = pl_block("SCENARIO B — 35,000 SF", "C", R["rev_total"], C["equip"], "C", r + 2)

put(ws, PL["A"]["rev"], 7, "Ramp applied to stabilised revenue", NOTE)
put(ws, PL["A"]["rent"], 7, "Fixed — does not scale with the ramp", NOTE)
put(ws, PL["A"]["ebitda"], 7, "Venue-level EBITDA, before corporate overhead", NOTE)
put(ws, PL["B"]["ebitda"], 7, "Venue-level EBITDA, before corporate overhead", NOTE)

# =============================================== RETURNS & SCENARIOS =========
ws = wb.create_sheet("Returns & Scenarios")
ws.sheet_view.showGridLines = False
for k, v in {"A": 46, "B": 16, "C": 16, "D": 52}.items():
    ws.column_dimensions[k].width = v

r = 1
put(ws, r, 1, "RETURNS & SCENARIOS", TITLE); r += 1
put(ws, r, 1, "Venue-level returns on total project cost, excluding real estate.", NOTE); r += 2
put(ws, r, 2, "18,000 SF", H2); put(ws, r, 3, "35,000 SF", H2); put(ws, r, 4, "Note", H2); r += 1

RT = {}


def trow(key, text, fa, fb, fmt, note_text="", bold=False, fill=None, top=False):
    global r
    label(ws, r, text, bold=bold)
    bd = TOPLINE if top else None
    put(ws, r, 2, fa, TOTAL if bold else BODY, fmt, fill, bd)
    put(ws, r, 3, fb, TOTAL if bold else BODY, fmt, fill, bd)
    if note_text:
        note(ws, r, 4, note_text)
    RT[key] = r
    r += 1


band(ws, r, "BASE CASE RETURNS"); r += 1
trow("capex", "Total project cost",
     f"=Capex!$B${C['total']}", f"=Capex!$C${C['total']}", MONEY)
for cc in (2, 3):
    ws.cell(row=RT['capex'], column=cc).font = LINK
trow("rev3", "Year 3 revenue (stabilised)",
     f"=P&L!D{PL['A']['rev']}", f"=P&L!D{PL['B']['rev']}", MONEY)
trow("ebitda3", "Year 3 EBITDA",
     f"=P&L!D{PL['A']['ebitda']}", f"=P&L!D{PL['B']['ebitda']}", MONEY)
trow("margin3", "Year 3 EBITDA margin",
     f"=P&L!D{PL['A']['margin']}", f"=P&L!D{PL['B']['margin']}", PCT)
trow("ebitda5", "Year 5 EBITDA",
     f"=P&L!F{PL['A']['ebitda']}", f"=P&L!F{PL['B']['ebitda']}", MONEY)
trow("cum5", "Cumulative 5-year EBITDA",
     f"=P&L!F{PL['A']['cum']}", f"=P&L!F{PL['B']['cum']}", MONEY)
trow("payback", "Simple payback on stabilised EBITDA (years)",
     f"=IF(B{RT['ebitda3']}<=0,\"n/a\",B{RT['capex']}/B{RT['ebitda3']})",
     f"=IF(C{RT['ebitda3']}<=0,\"n/a\",C{RT['capex']}/C{RT['ebitda3']})", YRS,
     "Capex divided by Year 3 EBITDA", bold=True, fill=SUB)
trow("coc3", "Year 3 cash-on-cash return",
     f"=B{RT['ebitda3']}/B{RT['capex']}", f"=C{RT['ebitda3']}/C{RT['capex']}", PCT, "", bold=True, fill=SUB)
trow("recovery", "5-year EBITDA as % of capex",
     f"=B{RT['cum5']}/B{RT['capex']}", f"=C{RT['cum5']}/C{RT['capex']}", PCT,
     "Above 100% means the venue has repaid its build within five years")
trow("cap_per_dollar", "Capex per $1 of stabilised revenue",
     f"=B{RT['capex']}/B{RT['rev3']}", f"=C{RT['capex']}/C{RT['rev3']}", MULT,
     "Lower is better. This is where the smaller box tends to win", bold=True)
r += 1

band(ws, r, "SENSITIVITY — REVENUE SCENARIOS"); r += 1
put(ws, r, 2, "Revenue", H2); put(ws, r, 3, "multiplier", H2); r += 1

SC_START = r
scen = [
    ("Downside", 0.70, "Utilisation and spend disappoint; closer to benchmark density"),
    ("Base", 1.00, "The model as built"),
    ("Upside", 1.25, "Strong utilisation, F&B and membership outperform"),
]
sc_rows = {}
for name, mult, desc in scen:
    label(ws, r, name, bold=True)
    put(ws, r, 2, mult, INPUT, MULT, KEY)
    note(ws, r, 4, desc)
    sc_rows[name] = r
    r += 1
r += 1

put(ws, r, 1, "18,000 SF", H2)
put(ws, r, 2, "Revenue", H2); put(ws, r, 3, "EBITDA", H2); put(ws, r, 4, "Payback (yrs)", H2)
ws.cell(row=r, column=4).font = H2
r += 1
for name, _, _ in scen:
    sr = sc_rows[name]
    label(ws, r, "  " + name, indent=1)
    put(ws, r, 2, f"=Returns_A_rev*$B${sr}".replace("Returns_A_rev", f"'Returns & Scenarios'!$B${RT['rev3']}"),
        BODY, MONEY)
    # EBITDA = revenue*mult less costs, where fixed costs do not scale
    put(ws, r, 3,
        f"=B{r}-(B{r}*({AS('labour','B')}+{AS('mktg','B')}+{AS('repairs','B')}+{AS('admin','B')})"
        f"+B{r}*'Revenue Build'!$B${R['mix_fb']}*{AS('fb_cogs','B')}"
        f"+{AS('sf','B')}*({AS('rent_psf','B')}+{AS('util_psf','B')}+{AS('ins_psf','B')})"
        f"+Capex!$B${C['equip']}*{AS('svc','B')}+{AS('software','B')})",
        BODY, MONEY)
    put(ws, r, 4, f"=IF(C{r}<=0,\"n/a\",$B${RT['capex']}/C{r})", BODY, YRS)
    r += 1
r += 1

put(ws, r, 1, "35,000 SF", H2)
put(ws, r, 2, "Revenue", H2); put(ws, r, 3, "EBITDA", H2); put(ws, r, 4, "Payback (yrs)", H2)
ws.cell(row=r, column=4).font = H2
r += 1
for name, _, _ in scen:
    sr = sc_rows[name]
    label(ws, r, "  " + name, indent=1)
    put(ws, r, 2, f"='Returns & Scenarios'!$C${RT['rev3']}*$B${sr}", BODY, MONEY)
    put(ws, r, 3,
        f"=B{r}-(B{r}*({AS('labour','C')}+{AS('mktg','C')}+{AS('repairs','C')}+{AS('admin','C')})"
        f"+B{r}*'Revenue Build'!$C${R['mix_fb']}*{AS('fb_cogs','C')}"
        f"+{AS('sf','C')}*({AS('rent_psf','C')}+{AS('util_psf','C')}+{AS('ins_psf','C')})"
        f"+Capex!$C${C['equip']}*{AS('svc','C')}+{AS('software','C')})",
        BODY, MONEY)
    put(ws, r, 4, f"=IF(C{r}<=0,\"n/a\",$C${RT['capex']}/C{r})", BODY, YRS)
    r += 1
r += 1

band(ws, r, "RECONCILIATION AGAINST THE KIX PLAN"); r += 1
put(ws, r, 2, "KIX plan", H2); put(ws, r, 3, "This model (35k)", H2); put(ws, r, 4, "Gap", H2); r += 1

recon = [
    ("Stabilised revenue", f"=Benchmark!$B${KIX_REV}", f"=$C${RT['rev3']}", MONEY),
    ("Stabilised EBITDA", f"=Benchmark!$B${KIX_EBITDA}", f"=$C${RT['ebitda3']}", MONEY),
    ("EBITDA margin", f"=Benchmark!$B${KIX_EBITDA}/Benchmark!$B${KIX_REV}", f"=$C${RT['margin3']}", PCT),
    ("Buildout / project cost", f"=Benchmark!$B${KIX_CAPEX}", f"=$C${RT['capex']}", MONEY),
    ("Revenue per SF", f"=Benchmark!$B${KIX_PSF}", f"='Revenue Build'!$C${R['psf']}", MONEY2),
    ("Payback (years)", f"=Benchmark!$B${KIX_PAY}", f"=$C${RT['payback']}", YRS),
]
for text, fplan, fmodel, fmt in recon:
    label(ws, r, text, bold=True)
    put(ws, r, 2, fplan, LINK, fmt)
    put(ws, r, 3, fmodel, TOTAL, fmt)
    put(ws, r, 4, f"=IF(B{r}=0,0,C{r}/B{r}-1)", TOTAL, PCT, WARN)
    r += 1
r += 1

put(ws, r, 1, "What has to be true to reach the plan's $5.0M", H2); r += 1
for line in [
    "Raise blended station utilisation well above 20%, or add stations, or push spend per visit materially higher.",
    "Reach roughly 3x the revenue density of the only published ESA entertainment comparable.",
    "Build a real F&B operation — at 15% of revenue, as the benchmark runs, the number is unreachable.",
    "None of these are impossible. All of them are currently assumed rather than evidenced.",
    "",
    "The cheapest way to replace assumption with evidence: run the presell campaign (plan section 25) before the raise,",
    "and spend a day at ESA's own FTY Lab in Poole counting throughput per station per hour.",
]:
    put(ws, r, 1, line, BODY if line else NOTE); r += 1

# ---------------------------------------------------------------- freeze -----
for name in ["Benchmark", "Assumptions", "Revenue Build", "Capex", "Returns & Scenarios"]:
    wb[name].freeze_panes = "B6"
wb["P&L"].freeze_panes = "B7"

wb.calculation.fullCalcOnLoad = True
wb.save("/home/user/lodgient/kix-model/KIX_Venue_Model.xlsx")
print("written")

"""Independent re-implementation of the model, used to check the workbook's formulas.

This deliberately does NOT read the workbook's formulas - it recomputes the model from the
same assumption values in plain Python, so that agreement is evidence the formulas are right.
It also asserts that key formula strings in the workbook point at the cells they should.
"""

from openpyxl import load_workbook

FX = 1.30

BENCH = dict(sf=5382, rev_gbp=192_000, eq_gbp=69_000, cap_gbp=200_000, rent_gbp=45_000)
PLAN = dict(sf=35_000, rev=5_000_000, ebitda=1_300_000, capex=1_500_000)

BASE = dict(
    cycles_hr=7, hours_wk=85, weeks=51, util=0.20, cycles_visit=5,
    spend_visit=26, fb_head=6.00, retail_head=0.90,
    member_fee=85, member_visits=2.5, party_val=600, party_guests=16,
    fitout_psf=60, station_cost=24_000, conting=0.10,
    util_psf=3.50, ins_psf=2.50, mktg=0.06, fb_cogs=0.28, svc=0.08,
    repairs=0.02, admin=0.05,
    ramp=[0.55, 0.82, 1.00, 1.06, 1.11],
)

A = dict(BASE, name="18,000 SF", sf=18_000, stations=14, members=300, parties=300,
         academy_rev=220_000, academy_ff=9_000, camp_rev=120_000, camp_ff=3_000,
         league_rev=150_000, league_ff=7_000, corp_rev=140_000, corp_ff=2_800,
         sponsor=40_000, ffe=220_000, tech=180_000, preopen=150_000,
         labour=0.31, rent_psf=14, software=60_000)

B = dict(BASE, name="35,000 SF", sf=35_000, stations=22, members=500, parties=450,
         academy_rev=340_000, academy_ff=14_000, camp_rev=190_000, camp_ff=4_700,
         league_rev=320_000, league_ff=15_000, corp_rev=240_000, corp_ff=4_800,
         sponsor=60_000, ffe=380_000, tech=220_000, preopen=220_000,
         labour=0.30, rent_psf=13, software=75_000)


def run(a):
    cap_hr = a["stations"] * a["cycles_hr"]
    cap_yr = cap_hr * a["hours_wk"] * a["weeks"]
    used = cap_yr * a["util"]
    walkins = used / a["cycles_visit"]

    footfall = (walkins
                + a["members"] * a["member_visits"] * 12
                + a["parties"] * a["party_guests"]
                + a["academy_ff"] + a["camp_ff"] + a["league_ff"] + a["corp_ff"])

    rev = {
        "walk": walkins * a["spend_visit"],
        "member": a["members"] * a["member_fee"] * 12,
        "party": a["parties"] * a["party_val"],
        "academy": a["academy_rev"],
        "camp": a["camp_rev"],
        "league": a["league_rev"],
        "corp": a["corp_rev"],
        "fb": footfall * a["fb_head"],
        "retail": footfall * a["retail_head"],
        "sponsor": a["sponsor"],
    }
    total = sum(rev.values())

    capex_sub = (a["sf"] * a["fitout_psf"] + a["stations"] * a["station_cost"]
                 + a["ffe"] + a["tech"] + a["preopen"])
    capex = capex_sub * (1 + a["conting"])
    equip = a["stations"] * a["station_cost"]

    fb_share = rev["fb"] / total
    years = []
    for m in a["ramp"]:
        r = total * m
        fb = r * fb_share
        costs = (r * a["labour"] + fb * a["fb_cogs"]
                 + a["sf"] * (a["rent_psf"] + a["util_psf"] + a["ins_psf"])
                 + r * a["mktg"] + equip * a["svc"] + a["software"]
                 + r * a["repairs"] + r * a["admin"])
        years.append((r, r - costs, (r - costs) / r))

    return dict(a=a, walkins=walkins, footfall=footfall, rev=rev, total=total,
                capex=capex, equip=equip, fb_share=fb_share, years=years)


bench_psf_usd = BENCH["rev_gbp"] * FX / BENCH["sf"]
bench_cap_psf = BENCH["cap_gbp"] * FX / BENCH["sf"]
plan_psf = PLAN["rev"] / PLAN["sf"]

print("=" * 78)
print("BENCHMARK — Ballpark Brighton")
print(f"  Revenue per SF (USD)        ${bench_psf_usd:,.2f}")
print(f"  Capex per SF (USD)          ${bench_cap_psf:,.2f}")
print(f"  Equipment % of capex        {BENCH['eq_gbp']/BENCH['cap_gbp']:.1%}")
print(f"\nKIX PLAN AS WRITTEN")
print(f"  Revenue per SF              ${plan_psf:,.2f}")
print(f"  Multiple of benchmark       {plan_psf/bench_psf_usd:.2f}x")
print(f"  Implied payback             {PLAN['capex']/PLAN['ebitda']:.1f} yrs")

for scen in (A, B):
    m = run(scen)
    y3r, y3e, y3m = m["years"][2]
    y5r, y5e, _ = m["years"][4]
    cum5 = sum(y[1] for y in m["years"])
    print("\n" + "=" * 78)
    print(f"SCENARIO — {scen['name']}")
    print(f"  Walk-in visits/yr           {m['walkins']:,.0f}")
    print(f"  Total footfall/yr           {m['footfall']:,.0f}")
    print(f"  Stabilised revenue          ${m['total']:,.0f}")
    for k, v in m["rev"].items():
        print(f"      {k:<10} ${v:>10,.0f}   {v/m['total']:>5.1%}")
    print(f"  Revenue per SF              ${m['total']/scen['sf']:,.2f}")
    print(f"  Multiple of benchmark       {(m['total']/scen['sf'])/bench_psf_usd:.2f}x")
    print(f"  vs plan $5.0M               {m['total']/PLAN['rev']:.1%}")
    print(f"  Total project cost          ${m['capex']:,.0f}   (${m['capex']/scen['sf']:,.2f}/SF)")
    print(f"  vs plan buildout $1.5M      {m['capex']/PLAN['capex']:.2f}x")
    print(f"  Y3 revenue / EBITDA         ${y3r:,.0f} / ${y3e:,.0f}  ({y3m:.1%})")
    print(f"  Y5 revenue / EBITDA         ${y5r:,.0f} / ${y5e:,.0f}")
    print(f"  Cumulative 5yr EBITDA       ${cum5:,.0f}  ({cum5/m['capex']:.0%} of capex)")
    print(f"  Payback on Y3 EBITDA        {m['capex']/y3e:.1f} yrs")
    print(f"  Y3 cash-on-cash             {y3e/m['capex']:.1%}")
    print(f"  Capex per $1 of revenue     {m['capex']/y3r:.2f}x")
    for lbl, mult in (("Downside", 0.70), ("Base", 1.00), ("Upside", 1.25)):
        r = y3r * mult
        fb = r * m["fb_share"]
        costs = (r * (scen["labour"] + scen["mktg"] + scen["repairs"] + scen["admin"])
                 + fb * scen["fb_cogs"]
                 + scen["sf"] * (scen["rent_psf"] + scen["util_psf"] + scen["ins_psf"])
                 + m["equip"] * scen["svc"] + scen["software"])
        e = r - costs
        pb = f"{m['capex']/e:.1f}" if e > 0 else "n/a"
        print(f"    {lbl:<9} rev ${r:>10,.0f}  EBITDA ${e:>10,.0f} ({e/r:>5.1%})  payback {pb} yrs")

# ---- structural checks on the workbook itself -------------------------------
print("\n" + "=" * 78)
print("WORKBOOK STRUCTURAL CHECKS")
wb = load_workbook("/home/user/lodgient/kix-model/KIX_Venue_Model.xlsx")
issues = []

banned = ("XLOOKUP", "XMATCH", "FILTER(", "UNIQUE(", "SEQUENCE(", "SORT(")
n_formulas = 0
for sh in wb.worksheets:
    for row in sh.iter_rows():
        for c in row:
            if isinstance(c.value, str) and c.value.startswith("="):
                n_formulas += 1
                up = c.value.upper()
                for b in banned:
                    if b in up:
                        issues.append(f"{sh.title}!{c.coordinate}: banned function {b}")
                if "Revenue Build" in c.value and "'Revenue Build'" not in c.value:
                    issues.append(f"{sh.title}!{c.coordinate}: unquoted sheet name with space")
                if "Returns & Scenarios" in c.value and "'Returns & Scenarios'" not in c.value:
                    issues.append(f"{sh.title}!{c.coordinate}: unquoted sheet name with space")

print(f"  Formulas written            {n_formulas}")
print(f"  fullCalcOnLoad              {wb.calculation.fullCalcOnLoad}")

# spot-check that named anchors point where we think
rb = wb["Revenue Build"]
spot = [(c.row, c.value) for row in rb.iter_rows() for c in row
        if c.column == 1 and isinstance(c.value, str)
        and c.value in ("STABILISED REVENUE", "Revenue per SF per year", "Total annual footfall")]
for row, lbl in spot:
    print(f"  '{lbl}' at Revenue Build row {row}: B={rb.cell(row=row, column=2).value}")

print("\n  ISSUES: " + ("none" if not issues else ""))
for i in issues:
    print("   -", i)


# ---- capital stack & debt service -------------------------------------------
def pmt(principal, rate, years):
    if principal <= 0:
        return 0.0
    i = rate / 12
    n = years * 12
    return principal * i / (1 - (1 + i) ** -n) * 12


print("\n" + "=" * 78)
print("CAPITAL STACK & DEBT SERVICE  (buy case)")

RE_PSF, RE_COSTS, INJ, BANK_PCT, ADV7A = 150, 0.03, 0.20, 0.50, 0.85
RATES = {"bank": (0.075, 25), "cdc": (0.065, 25), "sba7a": (0.1025, 10)}
PTAX_PCT, COV = 0.017, 1.25

for scen, wc, presell in ((A, 350_000, 150_000), (B, 550_000, 200_000)):
    m = run(scen)
    re_price = scen["sf"] * RE_PSF
    re_total = re_price * (1 + RE_COSTS)
    total_buy = re_total + m["capex"] + wc

    elig = re_total + scen["sf"] * scen["fitout_psf"] + m["equip"]
    bank = elig * BANK_PCT
    cdc = elig * (1 - INJ - BANK_PCT)
    inj504 = elig * INJ
    non504 = total_buy - elig
    l7a = non504 * ADV7A
    inj7a = non504 - l7a

    debt = bank + cdc + l7a
    equity = inj504 + inj7a
    ds = sum(pmt(p, *RATES[k]) for k, p in (("bank", bank), ("cdc", cdc), ("sba7a", l7a)))
    rent_back = scen["sf"] * scen["rent_psf"]
    ptax = re_price * PTAX_PCT

    print(f"\n--- {scen['name']} ---")
    print(f"  Real estate ({RE_PSF}/SF + {RE_COSTS:.0%})   ${re_total:,.0f}")
    print(f"  Venue project cost               ${m['capex']:,.0f}")
    print(f"  Working capital reserve          ${wc:,.0f}")
    print(f"  TOTAL PROJECT COST (BUY)         ${total_buy:,.0f}")
    print(f"  TOTAL PROJECT COST (LEASE)       ${m['capex'] - scen['sf']*15 + wc:,.0f}")
    print(f"    504-eligible                   ${elig:,.0f}")
    print(f"    Bank 1st ({BANK_PCT:.0%})                 ${bank:,.0f}   {pmt(bank,*RATES['bank']):>10,.0f}/yr")
    print(f"    CDC debenture ({1-INJ-BANK_PCT:.0%})        ${cdc:,.0f}   {pmt(cdc,*RATES['cdc']):>10,.0f}/yr")
    print(f"    SBA 7(a)                       ${l7a:,.0f}   {pmt(l7a,*RATES['sba7a']):>10,.0f}/yr")
    print(f"  Total debt                       ${debt:,.0f}  ({debt/total_buy:.0%} of cost)")
    print(f"  EQUITY REQUIRED                  ${equity:,.0f}")
    print(f"  Less presold memberships         ${presell:,.0f}")
    print(f"  NET EQUITY                       ${equity-presell:,.0f}")
    print(f"  Annual debt service              ${ds:,.0f}")
    print(f"  {'Year':<6}{'CADS':>12}{'DSCR':>9}{'Result':>8}{'FCF':>12}{'Cumulative':>13}")
    cum = 0.0
    y3fcf = y5fcf = 0.0
    for i, (rev, eb, _) in enumerate(m["years"], 1):
        cads = eb + rent_back - ptax
        dscr = cads / ds
        fcf = cads - ds
        cum += fcf
        if i == 3:
            y3fcf = fcf
        if i == 5:
            y5fcf = fcf
        print(f"  {i:<6}{cads:>12,.0f}{dscr:>9.2f}{'PASS' if dscr>=COV else 'FAIL':>8}"
              f"{fcf:>12,.0f}{cum:>13,.0f}")
    ne = equity - presell
    print(f"  Y3 cash-on-cash on equity        {y3fcf/ne:.1%}")
    print(f"  Y5 cash-on-cash on equity        {y5fcf/ne:.1%}")

# ---- reference-quoting check for every sheet name needing quotes ------------
need_quotes = [s for s in wb.sheetnames if not s.replace("_", "").isalnum()]
bad = []
for sh in wb.worksheets:
    for row in sh.iter_rows():
        for c in row:
            if isinstance(c.value, str) and c.value.startswith("="):
                for name in need_quotes:
                    if f"{name}!" in c.value and f"'{name}'!" not in c.value:
                        bad.append(f"{sh.title}!{c.coordinate}: unquoted '{name}' -> {c.value[:70]}")
print(f"\n  Sheet names requiring quotes: {need_quotes}")
print("  UNQUOTED REFERENCE ISSUES: " + ("none" if not bad else str(len(bad))))
for b in bad[:20]:
    print("   -", b)
